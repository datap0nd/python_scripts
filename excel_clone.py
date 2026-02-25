"""
Excel File Cloner — bypasses application-level file saves.

Approach 1: Duplicate unzipped template → populate XMLs with COM data → rezip
Approach 2: Read via COM object model → rebuild with openpyxl (fallback)

Output: %LOCALAPPDATA%\Temp\<source_filename>.xlsx
"""

import win32com.client
import zipfile
import os
import shutil
import sys
import xml.etree.ElementTree as ET
from pathlib import Path

# ─── Config ───────────────────────────────────────────────────────────────────

TEMP_DIR = os.path.join(os.environ["LOCALAPPDATA"], "Temp")

# *** SET THIS to the path of your unzipped sample xlsx folder ***
TEMPLATE_DIR = r"C:\Users\r.cunha\AppData\Local\Temp\xlsx_template"

WORK_DIR = os.path.join(TEMP_DIR, "_xlclone_work")


# ─── Approach 1: Template duplicate + COM data → rezip ───────────────────────

SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
ET.register_namespace("", SPREADSHEET_NS)
ET.register_namespace("r", "http://schemas.openxmlformats.org/officeDocument/2006/relationships")


def approach_1(wb, output_path):
    """Duplicate unzipped template, populate with COM data, rezip."""
    print("\n[Approach 1] Template + COM data")

    if not os.path.isdir(TEMPLATE_DIR):
        print(f"  FAILED — template folder not found: {TEMPLATE_DIR}")
        print("  Falling back to Approach 2.")
        return False

    # Duplicate template folder
    if os.path.exists(WORK_DIR):
        shutil.rmtree(WORK_DIR)
    shutil.copytree(TEMPLATE_DIR, WORK_DIR)
    print(f"  Duplicated template → {WORK_DIR}")

    # Read data from open Excel via COM (from memory, unencrypted)
    for si in range(1, wb.Sheets.Count + 1):
        ws = wb.Sheets(si)
        used = ws.UsedRange
        if used is None:
            continue

        start_row = used.Row
        start_col = used.Column
        num_rows = used.Rows.Count
        num_cols = used.Columns.Count
        print(f"  Sheet '{ws.Name}': {num_rows} rows x {num_cols} cols")

        # Bulk read values
        raw = used.Value
        if num_rows == 1 and num_cols == 1:
            values = [[raw]]
        elif num_rows == 1:
            values = [list(raw)]
        elif num_cols == 1:
            values = [[v] for (v,) in raw] if isinstance(raw[0], tuple) else [[v] for v in raw]
        else:
            values = [list(row) for row in raw]

        # Build sheet XML
        sheet_xml = _build_sheet_xml(values, start_row, start_col, ws)

        # Write into the template's worksheet file
        sheet_path = os.path.join(WORK_DIR, "xl", "worksheets", f"sheet{si}.xml")
        with open(sheet_path, "w", encoding="utf-8") as f:
            f.write('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n')
            f.write(sheet_xml)
        print(f"  Wrote {sheet_path}")

    # Build shared strings
    _build_shared_strings(wb, WORK_DIR)

    # Zip into output
    if os.path.exists(output_path):
        os.remove(output_path)

    with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for root, _, files in os.walk(WORK_DIR):
            for f in files:
                full_path = os.path.join(root, f)
                arcname = os.path.relpath(full_path, WORK_DIR)
                zout.write(full_path, arcname)

    # Cleanup work dir
    shutil.rmtree(WORK_DIR)

    print(f"  SUCCESS → {output_path}")
    return True


def _build_sheet_xml(values, start_row, start_col, ws):
    """Build worksheet XML from COM-read values."""
    ns = SPREADSHEET_NS
    root = ET.Element(f"{{{ns}}}worksheet")
    sheet_data = ET.SubElement(root, f"{{{ns}}}sheetData")

    for i, row_vals in enumerate(values):
        r = start_row + i
        row_el = ET.SubElement(sheet_data, f"{{{ns}}}row", r=str(r))

        for j, val in enumerate(row_vals):
            c = start_col + j
            col_letter = _col_letter(c)
            cell_ref = f"{col_letter}{r}"
            cell_el = ET.SubElement(row_el, f"{{{ns}}}c", r=cell_ref)

            if val is None:
                continue
            elif isinstance(val, str):
                cell_el.set("t", "inlineStr")
                is_el = ET.SubElement(cell_el, f"{{{ns}}}is")
                t_el = ET.SubElement(is_el, f"{{{ns}}}t")
                t_el.text = val
            elif isinstance(val, bool):
                cell_el.set("t", "b")
                v_el = ET.SubElement(cell_el, f"{{{ns}}}v")
                v_el.text = "1" if val else "0"
            else:
                # Numbers, dates (COM returns dates as floats)
                v_el = ET.SubElement(cell_el, f"{{{ns}}}v")
                v_el.text = str(val)

                # Preserve number format from COM
                try:
                    nf = ws.Cells(r, c).NumberFormat
                    if nf and "yy" in str(nf).lower() or "/" in str(nf):
                        # Date format — could map to style index if needed
                        pass
                except Exception:
                    pass

    return ET.tostring(root, encoding="unicode")


def _build_shared_strings(wb, work_dir):
    """Build sharedStrings.xml (empty — we use inline strings instead)."""
    ns = SPREADSHEET_NS
    root = ET.Element(f"{{{ns}}}sst", count="0", uniqueCount="0")
    sst_path = os.path.join(work_dir, "xl", "sharedStrings.xml")
    tree = ET.ElementTree(root)
    with open(sst_path, "wb") as f:
        tree.write(f, xml_declaration=True, encoding="UTF-8")


def _col_letter(col_num):
    """Convert 1-based column number to Excel letter (1→A, 27→AA)."""
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result


# ─── Approach 2: COM object model + openpyxl rebuild ─────────────────────────

def approach_2(wb, output_path):
    """Read all data + formatting via COM, rebuild from scratch with openpyxl."""
    print("\n[Approach 2] COM read + openpyxl rebuild")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    total_sheets = wb.Sheets.Count
    for si in range(1, total_sheets + 1):
        ws = wb.Sheets(si)
        out_ws = out_wb.create_sheet(title=ws.Name)
        print(f"\n  Sheet {si}/{total_sheets}: '{ws.Name}'")

        # Get used range
        used = ws.UsedRange
        if used is None:
            print("    (empty)")
            continue

        start_row = used.Row
        start_col = used.Column
        num_rows = used.Rows.Count
        num_cols = used.Columns.Count
        print(f"    Range: {num_rows} rows x {num_cols} cols")

        # ── Bulk read values (much faster than cell-by-cell) ──
        raw = used.Value
        if num_rows == 1 and num_cols == 1:
            values = [[raw]]
        elif num_rows == 1:
            values = [list(raw)]
        elif num_cols == 1:
            values = [[v] for (v,) in raw] if isinstance(raw[0], tuple) else [[v] for v in raw]
        else:
            values = [list(row) for row in raw]

        # ── Write values + copy formatting ──
        for i, row_data in enumerate(values):
            for j, val in enumerate(row_data):
                r = start_row + i
                c = start_col + j
                out_cell = out_ws.cell(row=r, column=c, value=val)

                try:
                    src = ws.Cells(r, c)
                    _copy_font(src, out_cell)
                    _copy_fill(src, out_cell)
                    _copy_alignment(src, out_cell)
                    _copy_number_format(src, out_cell)
                    _copy_borders(src, out_cell)
                except Exception:
                    pass

            # Progress for large sheets
            if num_rows > 100 and (i + 1) % 500 == 0:
                print(f"    ... {i + 1}/{num_rows} rows")

        # ── Column widths ──
        for c in range(start_col, start_col + num_cols):
            try:
                w = ws.Columns(c).ColumnWidth
                if w:
                    out_ws.column_dimensions[get_column_letter(c)].width = float(w)
            except Exception:
                pass

        # ── Row heights ──
        for r in range(start_row, start_row + num_rows):
            try:
                h = ws.Rows(r).RowHeight
                if h:
                    out_ws.row_dimensions[r].height = float(h)
            except Exception:
                pass

        # ── Merged cells ──
        merged_done = set()
        for i in range(num_rows):
            for j in range(num_cols):
                r = start_row + i
                c = start_col + j
                try:
                    cell = ws.Cells(r, c)
                    if cell.MergeCells:
                        addr = cell.MergeArea.Address.replace("$", "")
                        if addr not in merged_done:
                            merged_done.add(addr)
                            out_ws.merge_cells(addr)
                except Exception:
                    pass

        if merged_done:
            print(f"    Merged regions: {len(merged_done)}")

    out_wb.save(output_path)
    print(f"\n  SUCCESS → {output_path}")
    return True


# ─── Formatting helpers ──────────────────────────────────────────────────────

def _bgr_to_hex(color_int):
    """Convert Windows BGR color integer to hex RGB string."""
    c = int(color_int)
    b = (c >> 16) & 0xFF
    g = (c >> 8) & 0xFF
    r = c & 0xFF
    return f"{r:02X}{g:02X}{b:02X}"


def _copy_font(src, out_cell):
    from openpyxl.styles import Font
    sf = src.Font
    kwargs = {}
    if sf.Name:
        kwargs["name"] = sf.Name
    if sf.Size:
        kwargs["size"] = sf.Size
    if sf.Bold is not None:
        kwargs["bold"] = bool(sf.Bold)
    if sf.Italic is not None:
        kwargs["italic"] = bool(sf.Italic)
    if sf.Strikethrough is not None:
        kwargs["strike"] = bool(sf.Strikethrough)
    # Underline: 2 = xlUnderlineStyleSingle, -4142 = xlNone
    if sf.Underline and sf.Underline == 2:
        kwargs["underline"] = "single"
    elif sf.Underline and sf.Underline == 4:
        kwargs["underline"] = "double"
    try:
        if sf.Color and sf.Color.RGB and sf.Color.RGB != 0:
            rgb = sf.Color.RGB
            if isinstance(rgb, int):
                kwargs["color"] = _bgr_to_hex(rgb)
    except Exception:
        pass
    if kwargs:
        out_cell.font = Font(**kwargs)


def _copy_fill(src, out_cell):
    from openpyxl.styles import PatternFill
    interior = src.Interior
    # Pattern: -4142 = xlNone, 1 = xlSolid
    if interior.Pattern and interior.Pattern != -4142:
        try:
            rgb = _bgr_to_hex(interior.Color)
            out_cell.fill = PatternFill(
                start_color=rgb, end_color=rgb, fill_type="solid"
            )
        except Exception:
            pass


def _copy_alignment(src, out_cell):
    from openpyxl.styles import Alignment
    ha_map = {
        -4131: "left", -4108: "center", -4152: "right",
        -4130: "justify", 1: "general", 5: "fill", 7: "distributed",
    }
    va_map = {
        -4160: "top", -4108: "center", -4107: "bottom",
        -4130: "justify", 5: "distributed",
    }
    out_cell.alignment = Alignment(
        horizontal=ha_map.get(src.HorizontalAlignment, "general"),
        vertical=va_map.get(src.VerticalAlignment, "bottom"),
        wrap_text=bool(src.WrapText) if src.WrapText else False,
        text_rotation=int(src.Orientation) if src.Orientation else 0,
        indent=int(src.IndentLevel) if src.IndentLevel else 0,
    )


def _copy_number_format(src, out_cell):
    nf = src.NumberFormat
    if nf:
        out_cell.number_format = str(nf)


def _copy_borders(src, out_cell):
    from openpyxl.styles import Border, Side

    STYLE_MAP = {
        1: "thin", 2: "medium", 3: "dashed", 4: "dotted",
        5: "thick", 6: "double", 7: "hair",
        8: "mediumDashed", 9: "dashDot", 10: "mediumDashDot",
        11: "dashDotDot", 12: "mediumDashDotDot", 13: "slantDashDot",
    }
    EDGE_MAP = {
        "left": 7,    # xlEdgeLeft
        "right": 10,  # xlEdgeRight
        "top": 8,     # xlEdgeTop
        "bottom": 9,  # xlEdgeBottom
    }

    sides = {}
    for name, edge_idx in EDGE_MAP.items():
        try:
            border = src.Borders(edge_idx)
            style = STYLE_MAP.get(border.LineStyle)
            if style:
                kwargs = {"style": style}
                try:
                    if border.Color is not None:
                        kwargs["color"] = _bgr_to_hex(border.Color)
                except Exception:
                    pass
                sides[name] = Side(**kwargs)
        except Exception:
            pass

    if sides:
        out_cell.border = Border(**sides)


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    # Connect to running Excel
    try:
        excel = win32com.client.GetActiveObject("Excel.Application")
    except Exception as e:
        print(f"ERROR: Could not connect to Excel. Is it running?\n  {e}")
        sys.exit(1)

    count = excel.Workbooks.Count
    if count == 0:
        print("ERROR: No workbooks are open in Excel.")
        sys.exit(1)

    # List all open workbooks and let user pick
    workbooks = []
    for i in range(1, count + 1):
        workbooks.append(excel.Workbooks(i))

    if count == 1:
        wb = workbooks[0]
    else:
        print("\nOpen workbooks:")
        for i, w in enumerate(workbooks, 1):
            active = " ← (active)" if w.Name == excel.ActiveWorkbook.Name else ""
            print(f"  {i}. {w.Name}{active}")
        print(f"  0. All of them")

        choice = input("\nWhich one? [1]: ").strip()
        if choice == "0":
            for w in workbooks:
                _clone(w, os.path.join(TEMP_DIR, w.Name))
            print("\nDone! All files saved.")
            return
        elif choice == "":
            wb = workbooks[0]
        else:
            wb = workbooks[int(choice) - 1]

    source_name = wb.Name
    output_path = os.path.join(TEMP_DIR, source_name)

    _clone(wb, output_path)
    print(f"\nDone! File saved to:\n  {output_path}")


def _clone(wb, output_path):
    """Run Approach 1, fall back to Approach 2."""
    print(f"\nSource:  {wb.Name}")
    print(f"Output:  {output_path}")
    if not approach_1(wb, output_path):
        approach_2(wb, output_path)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
    input("\nPress Enter to exit...")
